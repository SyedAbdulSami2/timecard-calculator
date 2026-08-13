from __future__ import annotations

import base64
import csv
import io
import re
from pathlib import Path
from typing import Any

from PIL import Image

from .document_converter import convert_document


SUPPORTED = {
    ".pdf",
    ".xlsx",
    ".csv",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
}


HEADER_MAP = {
    "employee name": "employee_name",
    "employee": "employee_name",
    "week start": "week_start",
    "week end": "week_end",
    "date": "date",
    "day": "day",
    "clock in": "clock_in",
    "time in": "clock_in",
    "clock out": "clock_out",
    "time out": "clock_out",
    "break": "break_minutes",
    "break minutes": "break_minutes",
    "regular hours": "regular_hours",
    "hours": "regular_hours",
    "overtime hours": "overtime_hours",
    "ot": "overtime_hours",
    "holiday hours": "holiday_hours",
    "on-call hours": "on_call_hours",
    "on call": "on_call_hours",
    "call-back hours": "call_back_hours",
    "call back": "call_back_hours",
    "callback": "call_back_hours",
}


# ============================================================
# BASIC HELPERS
# ============================================================

def _num(
    value: Any,
) -> float:
    try:
        return float(
            str(value).strip() or 0
        )

    except Exception:
        return 0.0


def _break(
    value: Any,
) -> int:
    text = (
        str(value)
        .strip()
        .lower()
    )

    if not text:
        return 0

    time_match = re.fullmatch(
        r"(\d{1,2})\s*:\s*(\d{1,2})",
        text,
    )

    if time_match:
        hours = int(
            time_match.group(1)
        )

        minutes = int(
            time_match.group(2)
        )

        return (
            hours * 60
            + minutes
        )

    number_match = re.search(
        r"(\d+)",
        text,
    )

    if number_match:
        return int(
            number_match.group(1)
        )

    return 0


def _empty_result() -> dict[str, Any]:
    return {
        "employee_name": "",
        "week_start": "",
        "week_end": "",
        "rows": [],
        "warnings": [],
    }


# ============================================================
# CSV
# ============================================================

def extract_csv(
    data: bytes,
) -> dict[str, Any]:
    text = data.decode(
        "utf-8-sig",
        errors="replace",
    )

    reader = csv.DictReader(
        io.StringIO(
            text
        )
    )

    rows: list[
        dict[str, Any]
    ] = []

    for raw in reader:
        normalized = {
            HEADER_MAP.get(
                (
                    key
                    or ""
                )
                .strip()
                .lower(),
                (
                    key
                    or ""
                )
                .strip()
                .lower(),
            ): value
            for key, value
            in raw.items()
        }

        rows.append(
            {
                "date":
                    normalized.get(
                        "date",
                        "",
                    ),

                "day":
                    normalized.get(
                        "day",
                        "",
                    ),

                "clock_in":
                    normalized.get(
                        "clock_in",
                        "",
                    ),

                "clock_out":
                    normalized.get(
                        "clock_out",
                        "",
                    ),

                "break_minutes":
                    _break(
                        normalized.get(
                            "break_minutes",
                            0,
                        )
                    ),

                "regular_hours":
                    _num(
                        normalized.get(
                            "regular_hours",
                            0,
                        )
                    ),

                "overtime_hours":
                    _num(
                        normalized.get(
                            "overtime_hours",
                            0,
                        )
                    ),

                "holiday_hours":
                    _num(
                        normalized.get(
                            "holiday_hours",
                            0,
                        )
                    ),

                "on_call_hours":
                    _num(
                        normalized.get(
                            "on_call_hours",
                            0,
                        )
                    ),

                "call_back_hours":
                    _num(
                        normalized.get(
                            "call_back_hours",
                            0,
                        )
                    ),

                "total_hours":
                    0.0,
            }
        )

    return {
        "employee_name": "",
        "week_start": "",
        "week_end": "",
        "rows": rows,
        "warnings": [],
        "document_mode":
            "structured",
    }


# ============================================================
# XLSX
# ============================================================

def extract_xlsx(
    data: bytes,
) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        io.BytesIO(
            data
        ),
        data_only=True,
    )

    worksheet = (
        workbook.active
    )

    values = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    if not values:
        result = (
            _empty_result()
        )

        result[
            "document_mode"
        ] = "structured"

        return result

    headers = [
        str(
            value or ""
        )
        .strip()
        .lower()
        for value
        in values[0]
    ]

    keys = [
        HEADER_MAP.get(
            header,
            header,
        )
        for header
        in headers
    ]

    rows: list[
        dict[str, Any]
    ] = []

    for values_row in values[
        1:
    ]:
        if not any(
            value not in (
                None,
                "",
            )
            for value
            in values_row
        ):
            continue

        raw = dict(
            zip(
                keys,
                values_row,
            )
        )

        rows.append(
            {
                "date":
                    str(
                        raw.get(
                            "date",
                            "",
                        )
                        or ""
                    ),

                "day":
                    str(
                        raw.get(
                            "day",
                            "",
                        )
                        or ""
                    ),

                "clock_in":
                    str(
                        raw.get(
                            "clock_in",
                            "",
                        )
                        or ""
                    ),

                "clock_out":
                    str(
                        raw.get(
                            "clock_out",
                            "",
                        )
                        or ""
                    ),

                "break_minutes":
                    _break(
                        raw.get(
                            "break_minutes",
                            0,
                        )
                    ),

                "regular_hours":
                    _num(
                        raw.get(
                            "regular_hours",
                            0,
                        )
                    ),

                "overtime_hours":
                    _num(
                        raw.get(
                            "overtime_hours",
                            0,
                        )
                    ),

                "holiday_hours":
                    _num(
                        raw.get(
                            "holiday_hours",
                            0,
                        )
                    ),

                "on_call_hours":
                    _num(
                        raw.get(
                            "on_call_hours",
                            0,
                        )
                    ),

                "call_back_hours":
                    _num(
                        raw.get(
                            "call_back_hours",
                            0,
                        )
                    ),

                "total_hours":
                    0.0,
            }
        )

    return {
        "employee_name": "",
        "week_start": "",
        "week_end": "",
        "rows": rows,
        "warnings": [],
        "document_mode":
            "structured",
    }


# ============================================================
# DATA URL → PIL IMAGE
# ============================================================

def decode_data_image(
    value: str,
) -> Image.Image:
    """
    Convert:

    data:image/png;base64,...

    or:

    data:image/jpeg;base64,...

    back into a Pillow Image.
    """

    if not value:
        raise ValueError(
            "OCR image is empty."
        )

    if "," not in value:
        raise ValueError(
            "Invalid OCR image data."
        )

    _, encoded = (
        value.split(
            ",",
            1,
        )
    )

    try:
        raw = base64.b64decode(
            encoded
        )

        image = Image.open(
            io.BytesIO(
                raw
            )
        )

        image.load()

        return image.convert(
            "RGB"
        )

    except Exception as exc:
        raise ValueError(
            "Could not decode OCR image."
        ) from exc


# ============================================================
# BACKEND OCR
# ============================================================

def ocr_image(
    image: Image.Image,
    *,
    mode: str = "general",
) -> str:
    """
    OCR a prepared image on the backend.

    mode='general'
        allows labels/dates/day names.

    mode='times'
        concentrates on numerical/time values.
    """

    try:
        import pytesseract

    except ImportError as exc:
        raise RuntimeError(
            (
                "Backend OCR is not installed. "
                "Install pytesseract and the "
                "Tesseract OCR system package."
            )
        ) from exc

    if mode == "times":
        config = (
            "--oem 3 "
            "--psm 6 "
            "-c preserve_interword_spaces=1 "
            "-c tessedit_char_whitelist="
            "0123456789:/.-AMPampNnAa"
        )

    else:
        config = (
            "--oem 3 "
            "--psm 6 "
            "-c preserve_interword_spaces=1"
        )

    try:
        text = pytesseract.image_to_string(
            image,
            lang="eng",
            config=config,
        )

    except Exception as exc:
        raise RuntimeError(
            f"Backend OCR failed: {exc}"
        ) from exc

    return (
        text
        or ""
    ).strip()


# ============================================================
# OCR CONVERTED PAGE
# ============================================================

def ocr_converted_page(
    page: dict[str, Any],
) -> str:
    """
    OCR one converted page.

    We deliberately avoid OCRing every available image.

    Priority:
    1. table_image
    2. upper_table_image
    3. ocr_image
    4. full_ocr_image
    5. image

    This prevents the extremely slow repeated OCR behavior
    previously seen in the browser.
    """

    sources = [
        (
            "table",
            page.get(
                "table_image"
            ),
        ),
        (
            "upper_table",
            page.get(
                "upper_table_image"
            ),
        ),
        (
            "ocr",
            page.get(
                "ocr_image"
            ),
        ),
        (
            "full",
            page.get(
                "full_ocr_image"
            ),
        ),
        (
            "preview",
            page.get(
                "image"
            ),
        ),
    ]

    selected_name = ""
    selected_source = ""

    for (
        source_name,
        source_value,
    ) in sources:
        if source_value:
            selected_name = (
                source_name
            )

            selected_source = (
                source_value
            )

            break

    if not selected_source:
        return ""

    image = decode_data_image(
        selected_source
    )

    general_text = ocr_image(
        image,
        mode="general",
    )

    time_text = ocr_image(
        image,
        mode="times",
    )

    return "\n".join(
        [
            (
                f"OCR_SOURCE: "
                f"{selected_name}"
            ),
            general_text,
            "OCR_TIMES:",
            time_text,
        ]
    ).strip()


# ============================================================
# CONVERT + BACKEND OCR
# ============================================================

def extract_ocr_text(
    filename: str,
    data: bytes,
) -> dict[str, Any]:
    """
    Convert any supported PDF/image document and OCR it.

    Returns OCR text only.

    Structured parsing belongs in timecard_extractor.py.
    """

    converted = convert_document(
        filename,
        data,
    )

    pages = converted.get(
        "pages",
        [],
    )

    if not pages:
        raise ValueError(
            (
                "The document was converted, "
                "but no pages were produced."
            )
        )

    page_texts: list[str] = []

    for index, page in enumerate(
        pages
    ):
        try:
            text = ocr_converted_page(
                page
            )

            if text:
                page_texts.append(
                    "\n".join(
                        [
                            (
                                "===== PAGE "
                                f"{index + 1} "
                                "====="
                            ),
                            text,
                        ]
                    )
                )

        except Exception as exc:
            page_texts.append(
                (
                    f"===== PAGE {index + 1} =====\n"
                    f"OCR_ERROR: {exc}"
                )
            )

    return {
        "filename":
            converted.get(
                "filename",
                filename,
            ),

        "page_count":
            converted.get(
                "page_count",
                len(
                    pages
                ),
            ),

        "text":
            "\n\n".join(
                page_texts
            ).strip(),

        "document_mode":
            "backend_ocr",
    }


# ============================================================
# PDF / IMAGE LEGACY RESPONSE
# ============================================================

def extract_convertible_document(
    filename: str,
    data: bytes,
) -> dict[str, Any]:
    """
    Legacy /extract behavior.

    It now performs backend OCR instead of merely returning
    converted images.
    """

    ocr_result = extract_ocr_text(
        filename,
        data,
    )

    text = (
        ocr_result.get(
            "text",
            ""
        )
        or ""
    )

    return {
        "employee_name": "",
        "week_start": "",
        "week_end": "",
        "rows": [],
        "warnings": (
            []
            if text
            else [
                (
                    "The document was converted, "
                    "but no readable OCR text "
                    "was detected."
                )
            ]
        ),
        "ocr_status": (
            "complete"
            if text
            else "manual_review_required"
        ),
        "document_mode":
            "backend_ocr",

        "filename":
            ocr_result.get(
                "filename",
                filename,
            ),

        "page_count":
            ocr_result.get(
                "page_count",
                0,
            ),

        "raw_text":
            text,
    }


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def extract_document(
    filename: str,
    data: bytes,
) -> dict[str, Any]:
    extension = Path(
        filename or ""
    ).suffix.lower()

    if extension not in SUPPORTED:
        raise ValueError(
            (
                "Unsupported file type. "
                "Supported formats are PDF, CSV, XLSX, "
                "JPG, JPEG, PNG, WEBP, BMP, TIF, "
                "and TIFF."
            )
        )

    if not data:
        raise ValueError(
            "The uploaded file is empty."
        )

    if extension == ".csv":
        return extract_csv(
            data
        )

    if extension == ".xlsx":
        return extract_xlsx(
            data
        )

    if extension in {
        ".pdf",
        ".jpg",
        ".jpeg",
        ".png",
        ".webp",
        ".bmp",
        ".tif",
        ".tiff",
    }:
        return extract_convertible_document(
            filename,
            data,
        )

    raise ValueError(
        "Unsupported file type."
    )
